"""
Sistema de exportacion de datos a Excel y reportes
Genera reportes, recibos y backups en formato Excel
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import CARPETA_REPORTES, FORMATO_FECHA_EXPORT
from logger import registrar_operacion, registrar_error

# Crear carpeta de reportes si no existe
if not os.path.exists(CARPETA_REPORTES):
    os.makedirs(CARPETA_REPORTES)

class ExportadorExcel:
    """Clase para generar reportes en Excel"""
    
    def __init__(self):
        self.estilos = self._configurar_estilos()
    
    def _configurar_estilos(self):
        """Configura estilos reutilizables para Excel"""
        return {
            'titulo': Font(name='Arial', size=14, bold=True, color='FFFFFF'),
            'subtitulo': Font(name='Arial', size=11, bold=True),
            'encabezado': Font(name='Arial', size=10, bold=True, color='FFFFFF'),
            'normal': Font(name='Arial', size=10),
            'negrita': Font(name='Arial', size=10, bold=True),
            'fondo_titulo': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'fondo_encabezado': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'fondo_pagado': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'fondo_pendiente': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'fondo_parcial': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'borde': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alineacion_centro': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'alineacion_derecha': Alignment(horizontal='right', vertical='center'),
            'alineacion_izquierda': Alignment(horizontal='left', vertical='center', wrap_text=True)
        }
    
    def exportar_pagos(self, cooperaciones, nombre_archivo=None):
        """
        Exporta reporte completo de pagos a Excel
        
        Args:
            cooperaciones (list): Lista de cooperaciones con datos de pagos
            nombre_archivo (str): Nombre del archivo (opcional)
            
        Returns:
            str: Ruta del archivo creado o None si fallo
        """
        try:
            if nombre_archivo is None:
                fecha_actual = datetime.now().strftime('%d_%m_%Y_%H%M%S')
                nombre_archivo = f'Reporte_Pagos_{fecha_actual}.xlsx'
            
            ruta_archivo = os.path.join(CARPETA_REPORTES, nombre_archivo)
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Reporte de Pagos'
            
            # Encabezado
            ws.merge_cells('A1:H1')
            titulo = ws['A1']
            titulo.value = 'REPORTE DE PAGOS - COOPERACIONES COMUNITARIAS'
            titulo.font = self.estilos['titulo']
            titulo.fill = self.estilos['fondo_titulo']
            titulo.alignment = self.estilos['alineacion_centro']
            ws.row_dimensions[1].height = 25
            
            # Fecha de generacion
            ws.merge_cells('A2:H2')
            fecha = ws['A2']
            fecha.value = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
            fecha.font = self.estilos['subtitulo']
            fecha.alignment = self.estilos['alineacion_derecha']
            
            # Encabezados de tabla
            encabezados = ['Cooperacion', 'Proyecto', 'Monto Coop', 'Personas', 'Total Recaudado', 'Total Pendiente', 'Porcentaje', 'Estado']
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=4, column=col)
                celda.value = encabezado
                celda.font = self.estilos['encabezado']
                celda.fill = self.estilos['fondo_encabezado']
                celda.alignment = self.estilos['alineacion_centro']
                celda.border = self.estilos['borde']
            
            # Datos
            fila = 5
            total_recaudado = 0
            total_esperado = 0
            
            for coop in cooperaciones:
                nombre = coop.get('nombre', '')
                proyecto = coop.get('proyecto', '')
                monto_coop = float(coop.get('monto_cooperacion', 0))
                personas = coop.get('personas', [])
                
                # Calcular totales
                cantidad_personas = len(personas)
                recaudado = sum(sum(p.get('monto', 0) for p in persona.get('pagos', [])) 
                               for persona in personas)
                esperado = monto_coop * cantidad_personas
                pendiente = max(0, esperado - recaudado)
                porcentaje = (recaudado / esperado * 100) if esperado > 0 else 0
                
                total_recaudado += recaudado
                total_esperado += esperado
                
                # Determinar estado
                if pendiente == 0:
                    estado = 'COMPLETADO'
                    fondo = self.estilos['fondo_pagado']
                elif recaudado == 0:
                    estado = 'PENDIENTE'
                    fondo = self.estilos['fondo_pendiente']
                else:
                    estado = 'PARCIAL'
                    fondo = self.estilos['fondo_parcial']
                
                # Llenar fila
                valores = [nombre, proyecto, monto_coop, cantidad_personas, recaudado, pendiente, f'{porcentaje:.1f}%', estado]
                for col, valor in enumerate(valores, 1):
                    celda = ws.cell(row=fila, column=col)
                    celda.value = valor
                    celda.font = self.estilos['normal']
                    celda.fill = fondo
                    celda.border = self.estilos['borde']
                    
                    if col in [3, 5, 6]:  # Columnas numéricas
                        celda.alignment = self.estilos['alineacion_derecha']
                        if col == 3:
                            celda.number_format = '$#,##0.00'
                        elif col in [5, 6]:
                            celda.number_format = '$#,##0.00'
                    elif col == 7:  # Porcentaje
                        celda.alignment = self.estilos['alineacion_derecha']
                    else:
                        celda.alignment = self.estilos['alineacion_izquierda']
                
                fila += 1
            
            # Totales
            ws.merge_cells(f'A{fila}:D{fila}')
            total_label = ws[f'A{fila}']
            total_label.value = 'TOTALES'
            total_label.font = self.estilos['negrita']
            total_label.fill = self.estilos['fondo_titulo']
            total_label.alignment = self.estilos['alineacion_derecha']
            
            # Total recaudado
            celda_recaudado = ws[f'E{fila}']
            celda_recaudado.value = total_recaudado
            celda_recaudado.font = self.estilos['negrita']
            celda_recaudado.number_format = '$#,##0.00'
            celda_recaudado.fill = self.estilos['fondo_titulo']
            celda_recaudado.border = self.estilos['borde']
            
            # Total pendiente
            total_pendiente = total_esperado - total_recaudado
            celda_pendiente = ws[f'F{fila}']
            celda_pendiente.value = total_pendiente
            celda_pendiente.font = self.estilos['negrita']
            celda_pendiente.number_format = '$#,##0.00'
            celda_pendiente.fill = self.estilos['fondo_titulo']
            celda_pendiente.border = self.estilos['borde']
            
            # Porcentaje total
            porcentaje_total = (total_recaudado / total_esperado * 100) if total_esperado > 0 else 0
            celda_porcentaje = ws[f'G{fila}']
            celda_porcentaje.value = f'{porcentaje_total:.1f}%'
            celda_porcentaje.font = self.estilos['negrita']
            celda_porcentaje.fill = self.estilos['fondo_titulo']
            celda_porcentaje.border = self.estilos['borde']
            celda_porcentaje.alignment = self.estilos['alineacion_derecha']
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            
            wb.save(ruta_archivo)
            registrar_operacion('EXPORTACION', 'Exportar pagos a Excel', {'archivo': nombre_archivo})
            return ruta_archivo
            
        except Exception as e:
            registrar_error('exportador', 'exportar_pagos', str(e), f'cooperaciones: {len(cooperaciones)}')
            return None
    
    def exportar_personas_cooperacion(self, personas, nombre_cooperacion, nombre_archivo=None):
        """
        Exporta detalle de personas de una cooperacion
        
        Args:
            personas (list): Lista de personas
            nombre_cooperacion (str): Nombre de la cooperacion
            nombre_archivo (str): Nombre del archivo (opcional)
            
        Returns:
            str: Ruta del archivo creado o None si fallo
        """
        try:
            if nombre_archivo is None:
                fecha_actual = datetime.now().strftime('%d_%m_%Y_%H%M%S')
                nombre_archivo = f'Detalle_{nombre_cooperacion}_{fecha_actual}.xlsx'
            
            ruta_archivo = os.path.join(CARPETA_REPORTES, nombre_archivo)
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Detalle'
            
            # Encabezado
            ws.merge_cells('A1:F1')
            titulo = ws['A1']
            titulo.value = f'DETALLE DE PAGOS - {nombre_cooperacion.upper()}'
            titulo.font = self.estilos['titulo']
            titulo.fill = self.estilos['fondo_titulo']
            titulo.alignment = self.estilos['alineacion_centro']
            ws.row_dimensions[1].height = 25
            
            # Encabezados
            encabezados = ['Nombre', 'Folio', 'Monto Esperado', 'Total Pagado', 'Pendiente', 'Estado']
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=3, column=col)
                celda.value = encabezado
                celda.font = self.estilos['encabezado']
                celda.fill = self.estilos['fondo_encabezado']
                celda.alignment = self.estilos['alineacion_centro']
                celda.border = self.estilos['borde']
            
            # Datos
            fila = 4
            for persona in personas:
                nombre = persona.get('nombre', '')
                folio = persona.get('folio', '')
                monto_esperado = float(persona.get('monto_esperado', 0))
                pagos = persona.get('pagos', [])
                total_pagado = sum(p.get('monto', 0) for p in pagos)
                pendiente = max(0, monto_esperado - total_pagado)
                
                estado = 'PAGADO' if pendiente == 0 else ('PENDIENTE' if total_pagado == 0 else 'PARCIAL')
                fondo = self.estilos['fondo_pagado'] if pendiente == 0 else (self.estilos['fondo_pendiente'] if total_pagado == 0 else self.estilos['fondo_parcial'])
                
                valores = [nombre, folio, monto_esperado, total_pagado, pendiente, estado]
                for col, valor in enumerate(valores, 1):
                    celda = ws.cell(row=fila, column=col)
                    celda.value = valor
                    celda.font = self.estilos['normal']
                    celda.fill = fondo
                    celda.border = self.estilos['borde']
                    
                    if col in [3, 4, 5]:
                        celda.alignment = self.estilos['alineacion_derecha']
                        celda.number_format = '$#,##0.00'
                    else:
                        celda.alignment = self.estilos['alineacion_izquierda']
                
                fila += 1
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            
            wb.save(ruta_archivo)
            registrar_operacion('EXPORTACION', 'Exportar detalle cooperacion', {'cooperacion': nombre_cooperacion, 'archivo': nombre_archivo})
            return ruta_archivo
            
        except Exception as e:
            registrar_error('exportador', 'exportar_personas_cooperacion', str(e), f'cooperacion: {nombre_cooperacion}')
            return None

# Instancia global
exportador = ExportadorExcel()
