"""
Exportación y estadísticas del módulo de censo
"""

import subprocess
import sys
from tkinter import filedialog, messagebox
from datetime import datetime
import os

from src.core.logger import registrar_operacion


def exportar_excel(root, habitantes, tree):
    """
    Exporta la lista actual de habitantes a un archivo Excel
    
    Args:
        root: Ventana raíz de tkinter
        habitantes: Lista completa de habitantes
        tree: TreeView con los habitantes mostrados
    
    Returns:
        True si fue exitoso, False en caso contrario
    """
    try:
        # Intentar importar openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            respuesta = messagebox.askyesno(
                "Módulo requerido",
                "El módulo 'openpyxl' no está instalado.\n\n"
                "¿Desea instalarlo ahora? Esto puede tardar unos momentos."
            )
            if respuesta:
                messagebox.showinfo("Instalando", "Instalando openpyxl...\nPor favor espere.")
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("Éxito", "Instalación completada. Intente exportar nuevamente.")
            return False
        
        # Obtener lista actual (respetando filtros)
        habitantes_exportar = []
        for item_id in tree.get_children():
            valores = tree.item(item_id, 'values')
            folio = valores[0]
            habitante = next((h for h in habitantes if h['folio'] == folio), None)
            if habitante:
                habitantes_exportar.append(habitante)
        
        if not habitantes_exportar:
            messagebox.showwarning("Sin datos", "No hay habitantes para exportar")
            return False
        
        # Pedir ubicación de guardado
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"Censo_Habitantes_{fecha_actual}.xlsx"
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=nombre_archivo
        )
        
        if not archivo:
            return False
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Censo Habitantes"
        
        # Encabezados
        encabezados = ['Folio', 'Nombre Completo', 'Fecha de Registro', 'Estado', 'Nota']
        ws.append(encabezados)
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar datos
        for hab in habitantes_exportar:
            estado = "Activo" if hab.get('activo', True) else "Inactivo"
            fila = [
                hab.get('folio', ''),
                hab.get('nombre', ''),
                hab.get('fecha_registro', ''),
                estado,
                hab.get('nota', '')
            ]
            ws.append(fila)
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        
        # Guardar archivo
        wb.save(archivo)
        
        registrar_operacion('CENSO_EXPORTAR', f'Exportados {len(habitantes_exportar)} habitantes', 
                          {'archivo': archivo})
        
        messagebox.showinfo("Éxito", 
                          f"Se exportaron {len(habitantes_exportar)} habitantes correctamente.\n\n"
                          f"Archivo: {os.path.basename(archivo)}")
        return True
    
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar: {str(e)}")
        return False
