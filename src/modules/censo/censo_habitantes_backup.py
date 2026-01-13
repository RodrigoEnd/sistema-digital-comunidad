import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import requests
import subprocess
import sys
import time
import os
from datetime import datetime

# Configurar path para imports cuando se ejecuta directamente
if __name__ == "__main__":
    # Agregar la raíz del proyecto al path
    proyecto_raiz = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
    if proyecto_raiz not in sys.path:
        sys.path.insert(0, proyecto_raiz)

from src.config import API_URL, TEMAS
from src.ui.estilos_globales import TEMA_GLOBAL
from src.ui.tema_moderno import FUENTES
from src.core.logger import registrar_operacion, registrar_error
from src.core.validadores import validar_nombre
from src.modules.indicadores.indicadores_estado import calcular_estado_habitante

class SistemaCensoHabitantes:
    def __init__(self, root):
        self.root = root
        self.root.title("📋 Sistema de Censo de Habitantes - Comunidad San Pablo")
        self.root.state('zoomed')
        
        # Configurar icono y estilos
        self.root.geometry("1400x800")
        self.style = ttk.Style()
        
        self.api_url = API_URL
        self.habitantes = []
        self.nombre_visible = tk.BooleanVar(value=True)
        self.folio_visible = tk.BooleanVar(value=True)
        self.filtro_estado = tk.StringVar(value="Todos")  # Filtro de estado
        self._search_job = None  # Id de tarea programada para debounce
        
        # Variables para ordenamiento
        self.columna_orden = None
        self.orden_reversa = False
        
        # Cargar preferencias guardadas
        self.config_usuario_path = os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 
                                                'SistemaComunidad', 'config_censo_usuario.json')
        self.cargar_preferencias()
        
        # Referencias para las ventanas de procesos (evitar duplicados)
        self.proceso_control_pagos = None
        self.proceso_control_faenas = None
        
        # Estilos iniciales
        self.configurar_estilos()

        # Verificar/iniciar API local
        if not self.asegurar_api_activa():
            messagebox.showerror("Error", "No se pudo iniciar ni conectar con la API local")
            return
        
        self.configurar_interfaz()
        self.cargar_habitantes()

    def obtener_colores(self):
        return TEMA_GLOBAL
    
    def cargar_preferencias(self):
        """Carga las preferencias del usuario desde archivo JSON"""
        try:
            if os.path.exists(self.config_usuario_path):
                with open(self.config_usuario_path, 'r', encoding='utf-8') as f:
                    import json
                    prefs = json.load(f)
                    # Aplicar preferencias después de que se cree la interfaz
                    self.root.after(100, lambda: self.aplicar_preferencias(prefs))
        except Exception as e:
            print(f"Error cargando preferencias: {e}")
    
    def aplicar_preferencias(self, prefs):
        """Aplica las preferencias cargadas"""
        try:
            # Restaurar búsqueda
            if 'ultima_busqueda' in prefs:
                self.search_var.set(prefs['ultima_busqueda'])
            
            # Restaurar filtro de estado
            if 'filtro_estado' in prefs:
                self.filtro_estado.set(prefs['filtro_estado'])
            
            # Aplicar filtros
            self.aplicar_filtros()
        except Exception as e:
            print(f"Error aplicando preferencias: {e}")
    
    def guardar_preferencias(self):
        """Guarda las preferencias del usuario"""
        try:
            import json
            prefs = {
                'ultima_busqueda': self.search_var.get(),
                'filtro_estado': self.filtro_estado.get()
            }
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(self.config_usuario_path), exist_ok=True)
            
            with open(self.config_usuario_path, 'w', encoding='utf-8') as f:
                json.dump(prefs, f, indent=2)
        except Exception as e:
            print(f"Error guardando preferencias: {e}")

    def configurar_estilos(self):
        colores = self.obtener_colores()
        self.style.theme_use('clam')
        self.style.configure('TFrame', background=colores['bg_principal'])
        self.style.configure('TLabel', background=colores['bg_principal'], foreground=colores['fg_principal'])
        self.style.configure('TButton', background=colores['bg_secundario'], foreground=colores['fg_principal'], padding=6, borderwidth=1)
        self.style.map('TButton', background=[('active', colores['button_hover'])])
        self.style.configure('TEntry', fieldbackground=colores['input_bg'], borderwidth=1)
        self.style.configure('Treeview', background=colores['table_row_even'], fieldbackground=colores['table_row_even'], foreground=colores['fg_principal'], borderwidth=0)
        self.style.map('Treeview', background=[('selected', colores['table_selected'])], foreground=[('selected', colores['fg_principal'])])
        self.style.configure('Treeview.Heading', background=colores['table_header'], foreground=colores['table_header_text'], padding=6)

    def asegurar_api_activa(self):
        """Comprueba la API y la levanta si no está activa"""
        if self.verificar_api():
            return True
        
        # La API debe iniciarse desde main.py, no desde censo
        # Intentar iniciarla como última opción
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            proyecto_raiz = os.path.abspath(os.path.join(script_dir, "..", "..", ".."))
            api_path = os.path.join(proyecto_raiz, "src", "api", "api_local.py")
            
            if not os.path.exists(api_path):
                print(f"Error: No se encuentra API en {api_path}")
                return False
            
            # Lanzar la API en segundo plano
            subprocess.Popen([sys.executable, api_path], 
                           stdout=subprocess.DEVNULL, 
                           stderr=subprocess.DEVNULL,
                           creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
            
            # Esperar hasta 10 segundos a que responda
            print("Iniciando API local...")
            for i in range(20):
                time.sleep(0.5)
                if self.verificar_api():
                    print("API iniciada correctamente")
                    return True
            print("ADVERTENCIA: API tardó en iniciar")
            return True
        except Exception as e:
            print(f"Error iniciando API: {e}")
            return False
    
    def verificar_api(self):
        """Verificar que la API está funcionando"""
        try:
            response = requests.get(f"{self.api_url}/ping", timeout=2)
            return response.status_code == 200
        except:
            return False
    
    def configurar_interfaz(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        # La fila 1 contiene ahora el cuadro de lista+búsqueda
        main_frame.rowconfigure(1, weight=1)
        
        # ===== ENCABEZADO =====
        header_frame = ttk.LabelFrame(main_frame, text="Censo de Habitantes", padding="10")
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        # Guardar referencia para poder ocultarlo/mostrarlo al maximizar la lista
        self.header_frame = header_frame
        
        ttk.Label(header_frame, text="San Pablo Atotonilco - Sistema de Registro de Habitantes", 
               font=FUENTES['titulo']).grid(row=0, column=0, columnspan=3, pady=5)
        
        fecha_actual = datetime.now().strftime("%d/%m/%Y")
        ttk.Label(header_frame, text=f"Fecha: {fecha_actual}").grid(row=1, column=0, sticky=tk.W, padx=5)
        ttk.Button(header_frame, text="Abrir Control de Pagos", command=self.abrir_control_pagos, width=22).grid(row=1, column=1, padx=5)
        ttk.Button(header_frame, text="📊 Estadísticas", command=self.mostrar_estadisticas, width=18).grid(row=1, column=2, padx=5)
        
        self.total_label = ttk.Label(header_frame, text="Total Habitantes: 0", font=('Arial', 11, 'bold'))
        self.total_label.grid(row=1, column=3, sticky=tk.E, padx=5)

        # Indicadores de estado: pagos y faenas
        indicadores_frame = tk.Frame(header_frame, bg='#f0f0f0')
        indicadores_frame.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=5, pady=4)
        
        ttk.Label(indicadores_frame, text="Estado de la Comunidad:", font=FUENTES['pequeño']).pack(side=tk.LEFT, padx=5)
        
        # Indicador de pagos
        self.canvas_pagos = tk.Canvas(indicadores_frame, width=25, height=25, bg='#f0f0f0', 
                                       highlightthickness=0, cursor='hand2')
        self.canvas_pagos.pack(side=tk.LEFT, padx=2)
        self.canvas_pagos.bind('<Enter>', lambda e: self.mostrar_tooltip_pagos())
        self.canvas_pagos.bind('<Leave>', lambda e: self.ocultar_tooltip())
        
        self.label_pagos = ttk.Label(indicadores_frame, text="Pagos", font=FUENTES['pequeño'])
        self.label_pagos.pack(side=tk.LEFT, padx=(0, 10))
        
        # Indicador de faenas
        self.canvas_faenas = tk.Canvas(indicadores_frame, width=25, height=25, bg='#f0f0f0', 
                                        highlightthickness=0, cursor='hand2')
        self.canvas_faenas.pack(side=tk.LEFT, padx=2)
        self.canvas_faenas.bind('<Enter>', lambda e: self.mostrar_tooltip_faenas())
        self.canvas_faenas.bind('<Leave>', lambda e: self.ocultar_tooltip())
        
        self.label_faenas = ttk.Label(indicadores_frame, text="Faenas", font=FUENTES['pequeño'])
        self.label_faenas.pack(side=tk.LEFT)
        
        # Botón adicional: abrir faenas
        ttk.Button(header_frame, text="Abrir Registro de Faenas",
                    command=self.abrir_registro_faenas, width=24).grid(row=2, column=2, padx=5, pady=4, sticky=tk.E)
        
        # Tooltip
        self.tooltip_label = None
        
        # ===== CONTENEDOR PARA TABLA Y PANEL LATERAL =====
        contenedor_principal = ttk.Frame(main_frame)
        contenedor_principal.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        contenedor_principal.columnconfigure(0, weight=3)  # Tabla tiene más peso
        contenedor_principal.columnconfigure(1, weight=1)  # Panel lateral menos peso
        contenedor_principal.rowconfigure(0, weight=1)
        
        # ===== TABLA + BUSQUEDA EN EL MISMO CUADRO =====
        table_frame = ttk.LabelFrame(contenedor_principal, text="Listado de Habitantes", padding="8")
        table_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)

        # Toolbar de búsqueda y acciones dentro del mismo cuadro de la lista
        toolbar = ttk.Frame(table_frame)
        toolbar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        toolbar.columnconfigure(1, weight=1)

        # Fila 1: Búsqueda
        ttk.Label(toolbar, text="Buscar por nombre o folio:").grid(row=0, column=0, sticky=tk.W, padx=(0,5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.buscar_tiempo_real())
        search_entry = ttk.Entry(toolbar, textvariable=self.search_var, width=35)
        search_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))

        # Filtro de estado
        ttk.Label(toolbar, text="Filtrar:").grid(row=0, column=2, sticky=tk.W, padx=(10,5))
        filtro_combo = ttk.Combobox(toolbar, textvariable=self.filtro_estado, 
                                    values=["Todos", "Solo Activos", "Solo Inactivos"],
                                    state='readonly', width=15)
        filtro_combo.grid(row=0, column=3, sticky=tk.W, padx=(0,5))
        filtro_combo.bind('<<ComboboxSelected>>', lambda e: self.aplicar_filtros())

        # Fila 2: Botones de acción
        acciones = ttk.Frame(toolbar)
        acciones.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(8,0))
        ttk.Button(acciones, text="Limpiar", command=self.limpiar_busqueda).pack(side=tk.LEFT, padx=(0,5))
        ttk.Button(acciones, text="🔍 Búsqueda Avanzada", command=self.busqueda_avanzada).pack(side=tk.LEFT, padx=5)
        ttk.Button(acciones, text="Agregar Habitante", command=self.agregar_habitante).pack(side=tk.LEFT, padx=5)
        ttk.Button(acciones, text="Actualizar Lista", command=self.cargar_habitantes).pack(side=tk.LEFT, padx=5)
        ttk.Button(acciones, text="📊 Exportar a Excel", command=self.exportar_excel).pack(side=tk.LEFT, padx=5)

        # Fila 3: Controles de visibilidad y botón de tamaño
        controles = ttk.Frame(toolbar)
        controles.grid(row=2, column=0, columnspan=2, sticky=tk.E, pady=(6,0))
        ttk.Checkbutton(controles, text="Mostrar nombre", variable=self.nombre_visible,
                command=self.actualizar_visibilidad_columnas).pack(side=tk.RIGHT, padx=5)
        ttk.Checkbutton(controles, text="Mostrar folio", variable=self.folio_visible,
                command=self.actualizar_visibilidad_columnas).pack(side=tk.RIGHT, padx=5)
        ttk.Button(controles, text="Maximizar Lista", command=self.toggle_tamano_lista).pack(side=tk.RIGHT, padx=10)
        
        # Scrollbars
        scrollbar_y = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)
        scrollbar_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        
        # Treeview
        self.tree = ttk.Treeview(table_frame, 
                                 columns=('folio', 'nombre', 'fecha_registro', 'activo', 'nota'),
                                 show='headings',
                                 yscrollcommand=scrollbar_y.set,
                                 xscrollcommand=scrollbar_x.set)
        
        self.tree.heading('folio', text='Folio', command=lambda: self.ordenar_columna('folio'))
        self.tree.heading('nombre', text='Nombre Completo', command=lambda: self.ordenar_columna('nombre'))
        self.tree.heading('fecha_registro', text='Fecha de Registro', command=lambda: self.ordenar_columna('fecha_registro'))
        self.tree.heading('activo', text='Estado', command=lambda: self.ordenar_columna('activo'))
        self.tree.heading('nota', text='Nota', command=lambda: self.ordenar_columna('nota'))
        
        self.tree.column('folio', width=80, anchor=tk.CENTER, stretch=False)
        self.tree.column('nombre', width=280, anchor=tk.W)
        self.tree.column('fecha_registro', width=120, anchor=tk.CENTER)
        self.tree.column('activo', width=100, anchor=tk.CENTER)
        self.tree.column('nota', width=200, anchor=tk.W)
        
        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar_y.grid(row=1, column=1, sticky=(tk.N, tk.S))
        scrollbar_x.grid(row=2, column=0, sticky=(tk.W, tk.E))
        
        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)
        
        self.tree.tag_configure('activo', background='#c8e6c9')
        self.tree.tag_configure('inactivo', background='#ffccbc')

        # Menú contextual para acciones rápidas
        self.menu_contextual = tk.Menu(self.root, tearoff=0)
        self.menu_contextual.add_command(label="Marcar como activo", command=lambda: self.actualizar_estado_seleccion(True))
        self.menu_contextual.add_command(label="Marcar como inactivo", command=lambda: self.actualizar_estado_seleccion(False))
        self.menu_contextual.add_separator()
        self.menu_contextual.add_command(label="Colocar nota", command=self.colocar_nota_seleccion)
        self.tree.bind('<Button-3>', self.mostrar_menu_contextual)
        
        # Doble clic para ver/editar nota
        self.tree.bind('<Double-Button-1>', self.on_double_click)
        
        # Evento de selección para panel lateral
        self.tree.bind('<<TreeviewSelect>>', self.on_seleccionar_habitante)
        
        # Tooltip para notas truncadas
        self.tree.bind('<Motion>', self.on_hover_nota)
        self.tree.bind('<Leave>', lambda e: self.ocultar_tooltip_nota())
        self.tooltip_window = None
        
        # ===== PANEL LATERAL DE DETALLES =====
        self.panel_detalles = ttk.LabelFrame(contenedor_principal, text="Detalles del Habitante", padding="10")
        self.panel_detalles.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Mensaje inicial
        ttk.Label(self.panel_detalles, text="Selecciona un habitante\npara ver sus detalles",
                 font=('Arial', 10), foreground='#888', justify=tk.CENTER).pack(expand=True)
        
        self.habitante_seleccionado = None
        
        # Configurar atajos de teclado
        self.configurar_atajos_teclado()
        
        # Ajustar visibilidad inicial
        self.actualizar_visibilidad_columnas()

        # Estado de tamaño de lista (maximizada o normal)
        self.lista_maximizada = False
        
        # ===== BARRA DE ESTADO INFERIOR =====
        self.status_label = ttk.Label(main_frame, text="Cargando...", 
                                      font=('Arial', 9), foreground='#555')
        self.status_label.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(5, 0))
        self.status_label.config(relief=tk.SUNKEN, anchor=tk.W, padding=(5, 2))
    
    def configurar_atajos_teclado(self):
        """Configura atajos de teclado para acciones comunes"""
        # Ctrl+F: enfocar búsqueda
        self.root.bind('<Control-f>', lambda e: self.enfocar_busqueda())
        
        # Ctrl+N: agregar habitante
        self.root.bind('<Control-n>', lambda e: self.agregar_habitante())
        
        # F5: actualizar lista
        self.root.bind('<F5>', lambda e: self.cargar_habitantes())
        
        # Delete: marcar como inactivo
        self.root.bind('<Delete>', lambda e: self.actualizar_estado_seleccion(False))
        
        # Escape: limpiar búsqueda
        self.root.bind('<Escape>', lambda e: self.limpiar_busqueda())
    
    def enfocar_busqueda(self):
        """Coloca el foco en el campo de búsqueda"""
        # Buscar el Entry de búsqueda y darle foco
        for widget in self.root.winfo_children():
            if self._buscar_entry_recursivo(widget):
                return
    
    def _buscar_entry_recursivo(self, widget):
        """Busca recursivamente el Entry de búsqueda"""
        if isinstance(widget, ttk.Entry):
            try:
                if widget.cget('textvariable') == str(self.search_var):
                    widget.focus_set()
                    widget.select_range(0, tk.END)
                    return True
            except:
                pass
        for child in widget.winfo_children():
            if self._buscar_entry_recursivo(child):
                return True
        return False


    def toggle_tamano_lista(self):
        """Alterna entre vista normal y lista maximizada.
        En modo maximizado, se oculta el encabezado para dar más espacio
        a la lista; la barra de búsqueda se mantiene visible dentro del cuadro."""
        try:
            if not self.lista_maximizada:
                # Ocultar encabezado y dar más espacio a la lista
                if hasattr(self, 'header_frame'):
                    self.header_frame.grid_remove()
                # Asegurar que la ventana esté maximizada
                self.root.state('zoomed')
                self.lista_maximizada = True
            else:
                # Restaurar encabezado
                if hasattr(self, 'header_frame'):
                    self.header_frame.grid()
                self.lista_maximizada = False
        except Exception as e:
            print(f"Error al cambiar tamaño de lista: {e}")
    
    def cargar_habitantes(self):
        """Cargar todos los habitantes desde la API"""
        try:
            response = requests.get(f"{self.api_url}/habitantes", timeout=5)
            if response.status_code == 200:
                data = response.json()
                self.habitantes = data['habitantes']
                self.actualizar_tabla(self.habitantes)
            else:
                messagebox.showerror("Error", "No se pudieron cargar los habitantes")
        except Exception as e:
            messagebox.showerror("Error", f"Error de conexion: {str(e)}")
    
    def actualizar_tabla(self, habitantes):
        """Actualizar tabla con lista de habitantes"""
        # Limpiar tabla
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Agregar habitantes
        for hab in habitantes:
            activo = hab.get('activo', True)
            # Usar íconos visuales en lugar de texto
            estado_icono = "● Activo" if activo else "● Inactivo"
            tag = 'activo' if activo else 'inactivo'
            nota = hab.get('nota', '')
            
            self.tree.insert('', tk.END,
                           values=(hab['folio'],
                                  hab['nombre'],
                                  hab.get('fecha_registro', ''),
                                  estado_icono,
                                  nota[:30] + '...' if len(nota) > 30 else nota),
                           tags=(tag,))
        
        self.total_label.config(text=f"Total Habitantes: {len(habitantes)}")
        # Actualizar indicadores de estado
        self.actualizar_indicadores_estado()
        # Actualizar barra de estado inferior
        self.actualizar_barra_estado()

    def actualizar_barra_estado(self):
        """Actualiza la barra de estado inferior con información en tiempo real"""
        activos = sum(1 for hab in self.habitantes if hab.get('activo', True))
        inactivos = len(self.habitantes) - activos
        con_notas = sum(1 for hab in self.habitantes if hab.get('nota', ''))
        
        ultima_actualizacion = datetime.now().strftime("%H:%M:%S")
        
        texto_estado = f"📊 Total: {len(self.habitantes)} | ✓ Activos: {activos} | ✗ Inactivos: {inactivos} | 📝 Con notas: {con_notas} | Última actualización: {ultima_actualizacion}"
        
        if hasattr(self, 'status_label'):
            self.status_label.config(text=texto_estado)
    
    def ordenar_columna(self, columna):
        """Ordena la tabla por la columna especificada"""
        # Si es la misma columna, invertir orden; si es diferente, orden ascendente
        if self.columna_orden == columna:
            self.orden_reversa = not self.orden_reversa
        else:
            self.columna_orden = columna
            self.orden_reversa = False
        
        # Mapeo de columnas a claves en el diccionario
        mapeo_columnas = {
            'folio': 'folio',
            'nombre': 'nombre',
            'fecha_registro': 'fecha_registro',
            'activo': 'activo',
            'nota': 'nota'
        }
        
        clave = mapeo_columnas.get(columna)
        if not clave:
            return
        
        # Ordenar habitantes
        try:
            # Obtener lista actual mostrada (respetando filtros)
            lista_actual = []
            for item_id in self.tree.get_children():
                valores = self.tree.item(item_id, 'values')
                folio = valores[0]
                habitante = next((h for h in self.habitantes if h['folio'] == folio), None)
                if habitante:
                    lista_actual.append(habitante)
            
            # Ordenar según la columna
            if clave == 'activo':
                # Para activo, ordenar por booleano
                lista_actual.sort(key=lambda h: h.get(clave, True), reverse=self.orden_reversa)
            else:
                # Para las demás, ordenar por string/número
                lista_actual.sort(key=lambda h: str(h.get(clave, '')).lower(), reverse=self.orden_reversa)
            
            # Actualizar tabla con lista ordenada
            self.actualizar_tabla(lista_actual)
            
            # Actualizar indicador visual en encabezado
            for col in mapeo_columnas.keys():
                texto_base = {
                    'folio': 'Folio',
                    'nombre': 'Nombre Completo',
                    'fecha_registro': 'Fecha de Registro',
                    'activo': 'Estado',
                    'nota': 'Nota'
                }[col]
                
                if col == columna:
                    indicador = ' ▼' if self.orden_reversa else ' ▲'
                    self.tree.heading(col, text=texto_base + indicador)
                else:
                    self.tree.heading(col, text=texto_base)
        
        except Exception as e:
            print(f"Error al ordenar: {e}")


    def actualizar_indicadores_estado(self):
        """Actualiza los indicadores de estado de pagos y faenas para toda la comunidad"""
        try:
            total_pagos = 0.0
            total_faenas = 0.0
            count = 0
            
            for hab in self.habitantes:
                folio = hab.get('folio', '')
                nombre = hab.get('nombre', '')
                estado = calcular_estado_habitante(folio, nombre)
                total_pagos += estado['pagos']['ratio']
                total_faenas += estado['faenas']['ratio']
                count += 1
            
            if count > 0:
                ratio_pagos_promedio = total_pagos / count
                ratio_faenas_promedio = total_faenas / count
            else:
                ratio_pagos_promedio = 0.5
                ratio_faenas_promedio = 0.5
            
            # Dibujar rectángulos con color según ratio
            color_pagos = self._color_por_ratio(ratio_pagos_promedio)
            color_faenas = self._color_por_ratio(ratio_faenas_promedio)
            
            self.canvas_pagos.delete('all')
            self.canvas_pagos.create_rectangle(2, 2, 23, 23, fill=color_pagos, outline='#333', width=2)
            
            self.canvas_faenas.delete('all')
            self.canvas_faenas.create_rectangle(2, 2, 23, 23, fill=color_faenas, outline='#333', width=2)
            
            # Actualizar labels con porcentajes
            porcentaje_pagos = int(ratio_pagos_promedio * 100)
            porcentaje_faenas = int(ratio_faenas_promedio * 100)
            
            if hasattr(self, 'label_pagos'):
                self.label_pagos.config(text=f"Pagos: {porcentaje_pagos}%")
            if hasattr(self, 'label_faenas'):
                self.label_faenas.config(text=f"Faenas: {porcentaje_faenas}%")
            
            # Guardar para tooltip
            self.estado_pagos_promedio = ratio_pagos_promedio
            self.estado_faenas_promedio = ratio_faenas_promedio
        except Exception as e:
            print(f"Error actualizando indicadores: {e}")
    
    def _color_por_ratio(self, ratio):
        """Convierte ratio a color hexadecimal"""
        if ratio < 0.5:
            t = ratio * 2
            r = int(220 + (255 - 220) * t)
            g = int(53 + (193 - 53) * t)
            b = int(69 - 69 * t)
        else:
            t = (ratio - 0.5) * 2
            r = int(255 + (40 - 255) * t)
            g = int(193 + (167 - 193) * t)
            b = int(0 + (69 - 0) * t)
        return f"#{r:02x}{g:02x}{b:02x}"
    
    def mostrar_tooltip_pagos(self):
        """Muestra tooltip para indicador de pagos"""
        if hasattr(self, 'estado_pagos_promedio'):
            ratio = self.estado_pagos_promedio
            if ratio >= 0.8:
                texto = "Pagos: Al día ✓"
            elif ratio >= 0.5:
                texto = "Pagos: Deuda moderada"
            else:
                texto = "Pagos: Deuda crítica ⚠"
            self._mostrar_tooltip(texto)
    
    def mostrar_tooltip_faenas(self):
        """Muestra tooltip para indicador de faenas"""
        if hasattr(self, 'estado_faenas_promedio'):
            ratio = self.estado_faenas_promedio
            if ratio >= 0.8:
                texto = "Faenas: Comunidad muy activa"
            elif ratio >= 0.5:
                texto = "Faenas: Participación media"
            else:
                texto = "Faenas: Baja participación ⚠"
            self._mostrar_tooltip(texto)
    
    def _mostrar_tooltip(self, texto):
        """Muestra un tooltip flotante"""
        if self.tooltip_label:
            self.tooltip_label.destroy()
        self.tooltip_label = tk.Label(self.root, text=texto, bg='#333', fg='#fff', font=('Arial', 9),
                                      padx=8, pady=4, relief=tk.FLAT, borderwidth=1)
        self.tooltip_label.place(x=100, y=150)
    
    def on_hover_nota(self, event):
        """Muestra tooltip con nota completa cuando se pasa el mouse sobre una nota truncada"""
        # Identificar celda bajo el mouse
        item_id = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        
        # Solo mostrar tooltip para la columna de notas
        if column != '#5':  # Columna 5 es 'nota'
            self.ocultar_tooltip_nota()
            return
        
        if not item_id:
            self.ocultar_tooltip_nota()
            return
        
        # Obtener valores de la fila
        valores = self.tree.item(item_id, 'values')
        if len(valores) < 5:
            return
        
        nota_mostrada = valores[4]
        folio = valores[0]
        
        # Buscar nota completa del habitante
        habitante = next((h for h in self.habitantes if h['folio'] == folio), None)
        if not habitante:
            return
        
        nota_completa = habitante.get('nota', '')
        
        # Solo mostrar tooltip si la nota está truncada (más de 30 caracteres)
        if len(nota_completa) <= 30:
            self.ocultar_tooltip_nota()
            return
        
        # Crear tooltip si no existe o si cambió
        if self.tooltip_window:
            self.tooltip_window.destroy()
        
        # Crear ventana de tooltip
        self.tooltip_window = tk.Toplevel(self.root)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
        
        label = tk.Label(self.tooltip_window, text=nota_completa, 
                        background="#ffffcc", relief=tk.SOLID, borderwidth=1,
                        font=('Arial', 9), justify=tk.LEFT, padx=5, pady=3,
                        wraplength=300)
        label.pack()
    
    def ocultar_tooltip_nota(self):
        """Oculta el tooltip de nota"""
        if hasattr(self, 'tooltip_window') and self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None
    
    def ocultar_tooltip(self):
        """Oculta el tooltip"""
        if self.tooltip_label:
            self.tooltip_label.destroy()
            self.tooltip_label = None

    def actualizar_visibilidad_columnas(self):
        """Mostrar/ocultar columnas de nombre y folio"""
        # Evitar que ambas queden ocultas
        if not self.nombre_visible.get() and not self.folio_visible.get():
            self.folio_visible.set(True)
        
        if self.folio_visible.get():
            self.tree.column('folio', width=80, minwidth=40, stretch=False)
            self.tree.heading('folio', text='Folio')
        else:
            self.tree.column('folio', width=0, minwidth=0, stretch=False)
            self.tree.heading('folio', text='')
        
        if self.nombre_visible.get():
            self.tree.column('nombre', width=320, minwidth=120, stretch=True)
            self.tree.heading('nombre', text='Nombre Completo')
        else:
            self.tree.column('nombre', width=0, minwidth=0, stretch=False)
            self.tree.heading('nombre', text='')

    def mostrar_menu_contextual(self, event):
        """Selecciona fila bajo el cursor y muestra menú contextual"""
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
            self.menu_contextual.tk_popup(event.x_root, event.y_root)
        else:
            self.tree.selection_remove(self.tree.selection())

    def _folio_seleccionado(self):
        seleccion = self.tree.selection()
        if not seleccion:
            return None
        valores = self.tree.item(seleccion[0], 'values')
        if not valores:
            return None
        return valores[0]

    def actualizar_estado_seleccion(self, activo):
        """Marca habitante seleccionado como activo/inactivo"""
        folio = self._folio_seleccionado()
        if not folio:
            messagebox.showwarning("Seleccion", "Selecciona un habitante primero")
            return
        
        # Pedir confirmación solo al marcar como inactivo
        if not activo:
            habitante = next((h for h in self.habitantes if h['folio'] == folio), None)
            nombre = habitante['nombre'] if habitante else 'este habitante'
            respuesta = messagebox.askyesno(
                "Confirmar cambio", 
                f"¿Seguro que desea marcar como INACTIVO a {nombre}?\n\n"
                "Esta acción se puede revertir posteriormente.",
                icon='warning'
            )
            if not respuesta:
                return
        
        try:
            resp = requests.patch(f"{self.api_url}/habitantes/{folio}", json={'activo': activo}, timeout=5)
            if resp.status_code == 200:
                estado_txt = "activo" if activo else "inactivo"
                registrar_operacion('CENSO_ESTADO', f"Estado actualizado a {estado_txt}", {'folio': folio})
                self.cargar_habitantes()
            else:
                detalle = resp.text
                messagebox.showerror("Error", f"No se pudo actualizar el estado (HTTP {resp.status_code}): {detalle}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo actualizar: {e}")

    def colocar_nota_seleccion(self):
        """Permite agregar o editar nota del habitante seleccionado"""
        folio = self._folio_seleccionado()
        if not folio:
            messagebox.showwarning("Seleccion", "Selecciona un habitante primero")
            return
        nota = simpledialog.askstring("Nota", "Ingrese nota para el habitante:")
        if nota is None:
            return
        try:
            resp = requests.patch(f"{self.api_url}/habitantes/{folio}", json={'nota': nota}, timeout=5)
            if resp.status_code == 200:
                registrar_operacion('CENSO_NOTA', 'Nota actualizada', {'folio': folio})
                self.cargar_habitantes()
            else:
                detalle = resp.text
                messagebox.showerror("Error", f"No se pudo guardar la nota (HTTP {resp.status_code}): {detalle}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")
    
    def on_seleccionar_habitante(self, event=None):
        """Maneja la selección de un habitante para mostrar sus detalles"""
        folio = self._folio_seleccionado()
        if not folio:
            return
        
        habitante = next((h for h in self.habitantes if h['folio'] == folio), None)
        if not habitante:
            return
        
        self.habitante_seleccionado = habitante
        self.mostrar_panel_detalles(habitante)
    
    def mostrar_panel_detalles(self, habitante):
        """Muestra los detalles del habitante en el panel lateral"""
        # Limpiar panel
        for widget in self.panel_detalles.winfo_children():
            widget.destroy()
        
        # Título
        ttk.Label(self.panel_detalles, text="Información Completa", 
                 font=('Arial', 11, 'bold')).pack(pady=(0, 10), anchor=tk.W)
        
        # Datos del habitante
        datos_frame = ttk.Frame(self.panel_detalles)
        datos_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(datos_frame, text="Folio:", font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Label(datos_frame, text=habitante['folio'], font=('Arial', 9)).grid(row=0, column=1, sticky=tk.W, pady=2, padx=5)
        
        ttk.Label(datos_frame, text="Nombre:", font=('Arial', 9, 'bold')).grid(row=1, column=0, sticky=tk.W, pady=2)
        nombre_label = ttk.Label(datos_frame, text=habitante['nombre'], font=('Arial', 9), wraplength=180)
        nombre_label.grid(row=1, column=1, sticky=tk.W, pady=2, padx=5)
        
        ttk.Label(datos_frame, text="Fecha:", font=('Arial', 9, 'bold')).grid(row=2, column=0, sticky=tk.W, pady=2)
        ttk.Label(datos_frame, text=habitante.get('fecha_registro', 'N/A'), 
                 font=('Arial', 9)).grid(row=2, column=1, sticky=tk.W, pady=2, padx=5)
        
        activo = habitante.get('activo', True)
        estado_texto = "● Activo" if activo else "● Inactivo"
        estado_color = '#4CAF50' if activo else '#F44336'
        
        ttk.Label(datos_frame, text="Estado:", font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky=tk.W, pady=2)
        estado_label = ttk.Label(datos_frame, text=estado_texto, font=('Arial', 9), foreground=estado_color)
        estado_label.grid(row=3, column=1, sticky=tk.W, pady=2, padx=5)
        
        # Nota completa
        ttk.Separator(self.panel_detalles, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        ttk.Label(self.panel_detalles, text="Nota:", font=('Arial', 9, 'bold')).pack(anchor=tk.W, pady=(0, 5))
        
        nota_text = tk.Text(self.panel_detalles, height=5, width=25, font=('Arial', 9), 
                           wrap=tk.WORD, relief=tk.SOLID, borderwidth=1)
        nota_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        nota_text.insert('1.0', habitante.get('nota', 'Sin nota'))
        nota_text.config(state=tk.DISABLED)
        
        # Botones de acción
        ttk.Separator(self.panel_detalles, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=10)
        
        botones_frame = ttk.Frame(self.panel_detalles)
        botones_frame.pack(fill=tk.X)
        
        ttk.Button(botones_frame, text="Editar Nota", 
                  command=lambda: self.on_double_click(None)).pack(fill=tk.X, pady=2)
        ttk.Button(botones_frame, text="Ver Pagos", 
                  command=self.abrir_control_pagos).pack(fill=tk.X, pady=2)
        ttk.Button(botones_frame, text="Ver Faenas", 
                  command=self.abrir_registro_faenas).pack(fill=tk.X, pady=2)
        ttk.Button(botones_frame, text="Historial", 
                  command=lambda: self.mostrar_historial(habitante)).pack(fill=tk.X, pady=2)
        
        if activo:
            ttk.Button(botones_frame, text="Marcar Inactivo", 
                      command=lambda: self.actualizar_estado_seleccion(False)).pack(fill=tk.X, pady=2)
        else:
            ttk.Button(botones_frame, text="Marcar Activo", 
                      command=lambda: self.actualizar_estado_seleccion(True)).pack(fill=tk.X, pady=2)
    
    def mostrar_historial(self, habitante):
        """Muestra el historial de cambios del habitante"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Historial - {habitante['nombre']}")
        dialog.geometry("600x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text=f"Historial de: {habitante['nombre']}", 
                 font=('Arial', 11, 'bold')).pack(pady=10)
        ttk.Label(dialog, text=f"Folio: {habitante['folio']}", 
                 font=('Arial', 9)).pack(pady=5)
        
        # Frame para el texto del historial
        text_frame = ttk.Frame(dialog)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        historial_text = tk.Text(text_frame, font=('Courier', 9), wrap=tk.WORD,
                                yscrollcommand=scrollbar.set)
        historial_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=historial_text.yview)
        
        # Leer el log y filtrar por folio
        try:
            import json
            log_path = os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 
                                   'SistemaComunidad', 'sistema_comunidad.log')
            
            if os.path.exists(log_path):
                with open(log_path, 'r', encoding='utf-8') as f:
                    lineas = f.readlines()
                
                # Filtrar líneas que contengan el folio
                eventos = []
                for linea in lineas[-500:]:  # Últimas 500 líneas
                    if habitante['folio'] in linea or habitante['nombre'] in linea:
                        eventos.append(linea.strip())
                
                if eventos:
                    historial_text.insert('1.0', '\n'.join(eventos))
                else:
                    historial_text.insert('1.0', 'No se encontraron eventos registrados para este habitante.')
            else:
                historial_text.insert('1.0', 'No se encontró el archivo de log del sistema.')
        
        except Exception as e:
            historial_text.insert('1.0', f'Error al cargar historial: {str(e)}')
        
        historial_text.config(state=tk.DISABLED)
        
        ttk.Button(dialog, text="Cerrar", command=dialog.destroy).pack(pady=10)
    
    def mostrar_estadisticas(self):
        """Muestra estadísticas generales del censo"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Estadísticas del Censo")
        dialog.geometry("500x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Estadísticas del Censo", 
                 font=('Arial', 14, 'bold')).pack(pady=15)
        
        # Frame para estadísticas
        stats_frame = ttk.Frame(dialog, padding=20)
        stats_frame.pack(fill=tk.BOTH, expand=True)
        
        # Calcular estadísticas
        total = len(self.habitantes)
        activos = sum(1 for h in self.habitantes if h.get('activo', True))
        inactivos = total - activos
        con_notas = sum(1 for h in self.habitantes if h.get('nota', ''))
        
        # Estadísticas por fecha
        from collections import Counter
        fechas = [h.get('fecha_registro', '')[:7] for h in self.habitantes if h.get('fecha_registro')]  # YYYY-MM
        conteo_fechas = Counter(fechas)
        
        # Mostrar estadísticas generales
        ttk.Label(stats_frame, text="📊 GENERALES", font=('Arial', 11, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        
        stats_text = f"""
Total de habitantes: {total}
Activos: {activos} ({activos/total*100:.1f}%)
Inactivos: {inactivos} ({inactivos/total*100:.1f}%)
Con notas: {con_notas} ({con_notas/total*100:.1f}% si total > 0 else 0)
        """
        
        ttk.Label(stats_frame, text=stats_text, font=('Arial', 10), justify=tk.LEFT).pack(anchor=tk.W, pady=5)
        
        ttk.Separator(stats_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        # Registros por mes
        ttk.Label(stats_frame, text="📅 REGISTROS POR MES (últimos 6 meses)", 
                 font=('Arial', 11, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        
        if conteo_fechas:
            meses_ordenados = sorted(conteo_fechas.items(), reverse=True)[:6]
            for mes, cantidad in meses_ordenados:
                try:
                    mes_nombre = datetime.strptime(mes + "-01", "%Y-%m-%d").strftime("%B %Y")
                except:
                    mes_nombre = mes
                
                # Barra visual simple
                barra_largo = int(cantidad / max(conteo_fechas.values()) * 30)
                barra = "█" * barra_largo
                
                label_text = f"{mes_nombre}: {cantidad} habitantes {barra}"
                ttk.Label(stats_frame, text=label_text, font=('Courier', 9)).pack(anchor=tk.W, pady=2)
        else:
            ttk.Label(stats_frame, text="No hay datos de fechas disponibles", 
                     font=('Arial', 9), foreground='#888').pack(anchor=tk.W)
        
        ttk.Separator(stats_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        # Promedio de longitud de notas
        notas_con_contenido = [h.get('nota', '') for h in self.habitantes if h.get('nota', '')]
        if notas_con_contenido:
            promedio_longitud = sum(len(n) for n in notas_con_contenido) / len(notas_con_contenido)
            ttk.Label(stats_frame, text=f"📝 NOTAS", font=('Arial', 11, 'bold')).pack(anchor=tk.W, pady=(0, 5))
            ttk.Label(stats_frame, text=f"Promedio de caracteres por nota: {promedio_longitud:.0f}", 
                     font=('Arial', 10)).pack(anchor=tk.W)
        
        ttk.Button(dialog, text="Cerrar", command=dialog.destroy).pack(pady=15)
    
    def on_double_click(self, event):
        """Maneja el doble clic en la tabla para ver/editar nota"""
        # Si se llama desde el panel de detalles (sin evento), usar habitante seleccionado
        if event is None:
            if hasattr(self, 'habitante_seleccionado') and self.habitante_seleccionado:
                habitante = self.habitante_seleccionado
            else:
                folio = self._folio_seleccionado()
                if not folio:
                    return
                habitante = next((h for h in self.habitantes if h['folio'] == folio), None)
                if not habitante:
                    return
        else:
            # Identificar qué celda fue clickeada
            region = self.tree.identify_region(event.x, event.y)
            if region != "cell":
                return
            
            column = self.tree.identify_column(event.x)
            folio = self._folio_seleccionado()
            
            if not folio:
                return
            
            # Buscar habitante completo
            habitante = next((h for h in self.habitantes if h['folio'] == folio), None)
            if not habitante:
                return
        
        # Mostrar diálogo de nota completa
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Nota - {habitante['nombre']}")
        dialog.geometry("500x350")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text=f"Habitante: {habitante['nombre']}", 
                 font=('Arial', 10, 'bold')).pack(pady=10)
        ttk.Label(dialog, text=f"Folio: {habitante['folio']}", 
                 font=('Arial', 9)).pack(pady=5)
        
        ttk.Label(dialog, text="Nota:", font=('Arial', 10)).pack(pady=5, anchor=tk.W, padx=20)
        
        nota_text = tk.Text(dialog, width=55, height=10, font=('Arial', 10))
        nota_text.pack(pady=5, padx=20)
        nota_text.insert('1.0', habitante.get('nota', ''))
        
        def guardar_nota():
            nota_nueva = nota_text.get('1.0', 'end-1c').strip()
            try:
                resp = requests.patch(f"{self.api_url}/habitantes/{folio}", 
                                    json={'nota': nota_nueva}, timeout=5)
                if resp.status_code == 200:
                    registrar_operacion('CENSO_NOTA', 'Nota actualizada', {'folio': folio})
                    messagebox.showinfo("Éxito", "Nota guardada correctamente")
                    dialog.destroy()
                    self.cargar_habitantes()
                else:
                    messagebox.showerror("Error", f"HTTP {resp.status_code}: {resp.text}")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar: {e}")
        
        botones_frame = ttk.Frame(dialog)
        botones_frame.pack(pady=10)
        ttk.Button(botones_frame, text="Guardar", command=guardar_nota).pack(side=tk.LEFT, padx=5)
        ttk.Button(botones_frame, text="Cancelar", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def buscar_tiempo_real(self):
        """Busqueda en tiempo real mientras se escribe con debounce local"""
        # Cancelar una búsqueda pendiente para no saturar UI
        if self._search_job:
            self.root.after_cancel(self._search_job)
            self._search_job = None

        def _ejecutar_busqueda():
            self.aplicar_filtros()

        # Debounce corto para permitir escribir sin congelar la ventana
        self._search_job = self.root.after(180, _ejecutar_busqueda)
    
    def aplicar_filtros(self):
        """Aplica filtros de búsqueda y estado"""
        criterio = self.search_var.get().strip().lower()
        filtro = self.filtro_estado.get()
        
        # Filtrado local para evitar llamadas HTTP bloqueantes
        resultados = self.habitantes
        
        # Filtro de búsqueda por texto
        if criterio:
            resultados = [
                hab for hab in resultados
                if criterio in hab.get('nombre', '').lower()
                or criterio in hab.get('folio', '').lower()
            ]
        
        # Filtro de estado
        if filtro == "Solo Activos":
            resultados = [hab for hab in resultados if hab.get('activo', True)]
        elif filtro == "Solo Inactivos":
            resultados = [hab for hab in resultados if not hab.get('activo', True)]
        
        self.actualizar_tabla(resultados)
        
        # Guardar preferencias
        self.guardar_preferencias()
    
    def limpiar_busqueda(self):
        """Limpiar campo de busqueda y filtros"""
        self.search_var.set('')
        self.filtro_estado.set('Todos')
        self.actualizar_tabla(self.habitantes)
    
    def busqueda_avanzada(self):
        """Diálogo de búsqueda avanzada con múltiples filtros"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Búsqueda Avanzada")
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Búsqueda Avanzada", 
                 font=('Arial', 12, 'bold')).pack(pady=15)
        
        # Frame para filtros
        filtros_frame = ttk.LabelFrame(dialog, text="Criterios de Búsqueda", padding=15)
        filtros_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Nombre
        ttk.Label(filtros_frame, text="Nombre contiene:").grid(row=0, column=0, sticky=tk.W, pady=5)
        nombre_var = tk.StringVar()
        ttk.Entry(filtros_frame, textvariable=nombre_var, width=30).grid(row=0, column=1, pady=5, padx=10)
        
        # Estado
        ttk.Label(filtros_frame, text="Estado:").grid(row=1, column=0, sticky=tk.W, pady=5)
        estado_var = tk.StringVar(value="Todos")
        ttk.Combobox(filtros_frame, textvariable=estado_var, 
                    values=["Todos", "Activos", "Inactivos"], 
                    state='readonly', width=27).grid(row=1, column=1, pady=5, padx=10)
        
        # Rango de fechas
        ttk.Label(filtros_frame, text="Fecha desde (DD/MM/AAAA):").grid(row=2, column=0, sticky=tk.W, pady=5)
        fecha_desde_var = tk.StringVar()
        ttk.Entry(filtros_frame, textvariable=fecha_desde_var, width=30).grid(row=2, column=1, pady=5, padx=10)
        
        ttk.Label(filtros_frame, text="Fecha hasta (DD/MM/AAAA):").grid(row=3, column=0, sticky=tk.W, pady=5)
        fecha_hasta_var = tk.StringVar()
        ttk.Entry(filtros_frame, textvariable=fecha_hasta_var, width=30).grid(row=3, column=1, pady=5, padx=10)
        
        # Nota contiene
        ttk.Label(filtros_frame, text="Nota contiene:").grid(row=4, column=0, sticky=tk.W, pady=5)
        nota_var = tk.StringVar()
        ttk.Entry(filtros_frame, textvariable=nota_var, width=30).grid(row=4, column=1, pady=5, padx=10)
        
        # Con/sin notas
        tiene_nota_var = tk.StringVar(value="Todos")
        ttk.Label(filtros_frame, text="Tiene nota:").grid(row=5, column=0, sticky=tk.W, pady=5)
        ttk.Combobox(filtros_frame, textvariable=tiene_nota_var,
                    values=["Todos", "Con nota", "Sin nota"],
                    state='readonly', width=27).grid(row=5, column=1, pady=5, padx=10)
        
        def aplicar_busqueda():
            """Aplica todos los filtros de búsqueda avanzada"""
            resultados = self.habitantes
            
            # Filtro por nombre
            nombre = nombre_var.get().strip().lower()
            if nombre:
                resultados = [h for h in resultados if nombre in h.get('nombre', '').lower()]
            
            # Filtro por estado
            if estado_var.get() == "Activos":
                resultados = [h for h in resultados if h.get('activo', True)]
            elif estado_var.get() == "Inactivos":
                resultados = [h for h in resultados if not h.get('activo', True)]
            
            # Filtro por rango de fechas
            fecha_desde = fecha_desde_var.get().strip()
            fecha_hasta = fecha_hasta_var.get().strip()
            
            if fecha_desde or fecha_hasta:
                try:
                    if fecha_desde:
                        desde_dt = datetime.strptime(fecha_desde, "%d/%m/%Y")
                    if fecha_hasta:
                        hasta_dt = datetime.strptime(fecha_hasta, "%d/%m/%Y")
                    
                    resultados_filtrados = []
                    for h in resultados:
                        fecha_registro = h.get('fecha_registro', '')
                        if fecha_registro:
                            try:
                                fecha_dt = datetime.strptime(fecha_registro, "%d/%m/%Y")
                                incluir = True
                                if fecha_desde and fecha_dt < desde_dt:
                                    incluir = False
                                if fecha_hasta and fecha_dt > hasta_dt:
                                    incluir = False
                                if incluir:
                                    resultados_filtrados.append(h)
                            except:
                                pass
                    resultados = resultados_filtrados
                except ValueError:
                    messagebox.showerror("Error", "Formato de fecha inválido. Use DD/MM/AAAA")
                    return
            
            # Filtro por contenido de nota
            nota_texto = nota_var.get().strip().lower()
            if nota_texto:
                resultados = [h for h in resultados if nota_texto in h.get('nota', '').lower()]
            
            # Filtro por tiene/no tiene nota
            if tiene_nota_var.get() == "Con nota":
                resultados = [h for h in resultados if h.get('nota', '')]
            elif tiene_nota_var.get() == "Sin nota":
                resultados = [h for h in resultados if not h.get('nota', '')]
            
            # Mostrar resultados
            self.actualizar_tabla(resultados)
            messagebox.showinfo("Búsqueda Avanzada", 
                              f"Se encontraron {len(resultados)} habitantes que coinciden con los criterios.")
            dialog.destroy()
        
        # Botones
        botones_frame = ttk.Frame(dialog)
        botones_frame.pack(pady=15)
        ttk.Button(botones_frame, text="Buscar", command=aplicar_busqueda).pack(side=tk.LEFT, padx=5)
        ttk.Button(botones_frame, text="Cancelar", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def exportar_excel(self):
        """Exporta la lista actual de habitantes a un archivo Excel"""
        try:
            # Intentar importar openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                respuesta = messagebox.askyesno(
                    "Módulo requerido",
                    "El módulo 'openpyxl' no está instalado.\n\n"
                    "¿Desea instalarlo ahora? Esto puede tardar unos momentos."
                )
                if respuesta:
                    messagebox.showinfo("Instalando", "Instalando openpyxl...\nPor favor espere.")
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                    messagebox.showinfo("Éxito", "Instalación completada. Intente exportar nuevamente.")
                return
            
            # Obtener lista actual (respetando filtros)
            habitantes_exportar = []
            for item_id in self.tree.get_children():
                valores = self.tree.item(item_id, 'values')
                folio = valores[0]
                habitante = next((h for h in self.habitantes if h['folio'] == folio), None)
                if habitante:
                    habitantes_exportar.append(habitante)
            
            if not habitantes_exportar:
                messagebox.showwarning("Sin datos", "No hay habitantes para exportar")
                return
            
            # Pedir ubicación de guardado
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"Censo_Habitantes_{fecha_actual}.xlsx"
            archivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=nombre_archivo
            )
            
            if not archivo:
                return
            
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Censo Habitantes"
            
            # Encabezados
            encabezados = ['Folio', 'Nombre Completo', 'Fecha de Registro', 'Estado', 'Nota']
            ws.append(encabezados)
            
            # Estilo para encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Agregar datos
            for hab in habitantes_exportar:
                estado = "Activo" if hab.get('activo', True) else "Inactivo"
                fila = [
                    hab.get('folio', ''),
                    hab.get('nombre', ''),
                    hab.get('fecha_registro', ''),
                    estado,
                    hab.get('nota', '')
                ]
                ws.append(fila)
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 40
            
            # Guardar archivo
            wb.save(archivo)
            
            registrar_operacion('CENSO_EXPORTAR', f'Exportados {len(habitantes_exportar)} habitantes', 
                              {'archivo': archivo})
            
            messagebox.showinfo("Éxito", 
                              f"Se exportaron {len(habitantes_exportar)} habitantes correctamente.\n\n"
                              f"Archivo: {os.path.basename(archivo)}")
        
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {str(e)}")
    
    def buscar_nombres_similares(self, nombre):
        """Busca habitantes con nombres similares para evitar duplicados"""
        nombre_lower = nombre.lower()
        similares = []
        
        # Dividir nombre en palabras
        palabras_busqueda = nombre_lower.split()
        
        for hab in self.habitantes:
            nombre_hab = hab.get('nombre', '').lower()
            
            # Coincidencia exacta
            if nombre_hab == nombre_lower:
                similares.append(hab)
                continue
            
            # Coincidencia por palabras (al menos 2 palabras en común)
            palabras_hab = nombre_hab.split()
            palabras_comunes = set(palabras_busqueda) & set(palabras_hab)
            
            if len(palabras_comunes) >= min(2, len(palabras_busqueda)):
                similares.append(hab)
        
        return similares[:5]  # Limitar a 5 resultados
    
    def agregar_habitante(self):
        """Agregar nuevo habitante"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Agregar Nuevo Habitante")
        dialog.geometry("450x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text="Nombre Completo del Habitante:", 
                 font=('Arial', 10, 'bold')).pack(pady=10)
        nombre_entry = ttk.Entry(dialog, width=50)
        nombre_entry.pack(pady=10)
        
        def guardar(event=None):
            nombre = nombre_entry.get().strip()
            
            # Validar nombre
            try:
                nombre_validado = validar_nombre(nombre)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                return
            
            # Buscar nombres similares
            similares = self.buscar_nombres_similares(nombre_validado)
            if similares:
                nombres_str = '\n'.join([f"- {h['nombre']} (Folio: {h['folio']})" for h in similares])
                respuesta = messagebox.askyesno(
                    "Posible duplicado",
                    f"Se encontraron habitantes con nombres similares:\n\n{nombres_str}\n\n"
                    "¿Desea continuar agregando este habitante de todos modos?",
                    icon='warning'
                )
                if not respuesta:
                    return
            
            try:
                response = requests.post(f"{self.api_url}/habitantes",
                                        json={'nombre': nombre_validado},
                                        timeout=5)
                data = response.json()
                
                if data['success']:
                    registrar_operacion('CENSO_AGREGAR', 'Habitante agregado desde censo', 
                        {'nombre': nombre, 'folio': data['habitante']['folio']})
                    messagebox.showinfo("Exito", 
                        f"Habitante agregado correctamente\n"
                        f"Folio asignado: {data['habitante']['folio']}")
                    dialog.destroy()
                    self.cargar_habitantes()
                else:
                    registrar_error('CENSO_AGREGAR', data['mensaje'])
                    messagebox.showerror("Error", data['mensaje'])
            except Exception as e:
                messagebox.showerror("Error", f"Error al agregar: {str(e)}")
        
        botones = ttk.Frame(dialog)
        botones.pack(pady=10)
        ttk.Button(botones, text="Confirmar", command=guardar, width=18).pack(side=tk.LEFT, padx=5)
        ttk.Button(botones, text="Cancelar", command=dialog.destroy, width=12).pack(side=tk.LEFT, padx=5)
        dialog.bind("<Return>", guardar)
        nombre_entry.focus()

    def abrir_control_pagos(self):
        """Lanza la app de control de pagos - evita abrir duplicados"""
        try:
            # Verificar si ya existe un proceso abierto
            if self.proceso_control_pagos is not None:
                # Verificar si el proceso sigue ejecutándose
                if self.proceso_control_pagos.poll() is None:
                    # El proceso sigue activo, traerlo al frente
                    messagebox.showinfo("Información", "Control de Pagos ya está abierto.")
                    return
                else:
                    # El proceso terminó, permitir abrir uno nuevo
                    self.proceso_control_pagos = None
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            # Navegar a la carpeta de pagos (subir un nivel y entrar a pagos)
            pagos_path = os.path.join(script_dir, "..", "pagos", "control_pagos.py")
            pagos_path = os.path.abspath(pagos_path)  # Resolver la ruta absoluta
            
            if not os.path.exists(pagos_path):
                messagebox.showerror("Error", f"No se encontró el archivo:\n{pagos_path}")
                return
            
            # Abrir el sistema de control de pagos
            self.proceso_control_pagos = subprocess.Popen([sys.executable, pagos_path])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir Control de Pagos:\n{str(e)}")

    def abrir_registro_faenas(self):
        """Lanza la app de registro de faenas - evita abrir duplicados"""
        try:
            # Verificar si ya existe un proceso abierto
            if self.proceso_control_faenas is not None:
                # Verificar si el proceso sigue ejecutándose
                if self.proceso_control_faenas.poll() is None:
                    # El proceso sigue activo, mostrar mensaje
                    messagebox.showinfo("Información", "Registro de Faenas ya está abierto.")
                    return
                else:
                    # El proceso terminó, permitir abrir uno nuevo
                    self.proceso_control_faenas = None
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            # Navegar a la carpeta de faenas (subir un nivel y entrar a faenas)
            faenas_path = os.path.join(script_dir, "..", "faenas", "control_faenas.py")
            faenas_path = os.path.abspath(faenas_path)  # Resolver la ruta absoluta

            if not os.path.exists(faenas_path):
                messagebox.showerror("Error", f"No se encontró el archivo:\n{faenas_path}")
                return

            # Abrir con año por defecto 2025 para ver el resumen anual simulado
            self.proceso_control_faenas = subprocess.Popen([sys.executable, faenas_path, "--anio", "2025"])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir Registro de Faenas:\n{str(e)}")


def main():
    root = tk.Tk()
    app = SistemaCensoHabitantes(root)
    root.mainloop()

if __name__ == "__main__":
    main()
