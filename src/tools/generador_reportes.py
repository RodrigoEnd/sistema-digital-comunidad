"""
Generador Automatico de Reportes
Crea reportes en Excel y PDF con datos del sistema
"""

import os
from datetime import datetime
from src.config import CARPETA_REPORTES


class GeneradorReportes:
    """Clase para generar reportes del sistema"""
    
    def __init__(self):
        """Inicializar generador de reportes"""
        self.carpeta_reportes = CARPETA_REPORTES
        self._crear_carpeta_reportes()
    
    def _crear_carpeta_reportes(self):
        """Crear carpeta de reportes si no existe"""
        if not os.path.exists(self.carpeta_reportes):
            os.makedirs(self.carpeta_reportes)
    
    def _obtener_timestamp(self):
        """Obtener timestamp para nombres de archivo"""
        return datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # ========================================================================
    # REPORTES DE CENSO
    # ========================================================================
    
    def generar_reporte_censo_excel(self, habitantes, datos_usuario="Sistema"):
        """
        Genera reporte de censo en Excel
        
        Args:
            habitantes (list): Lista de habitantes
            datos_usuario (str): Usuario que genera el reporte
            
        Returns:
            str: Ruta del archivo generado
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Censo"
            
            # Titulo
            ws['A1'] = "REPORTE DE CENSO"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws['A3'] = f"Usuario: {datos_usuario}"
            
            # Encabezados
            encabezados = ['Folio', 'Nombre', 'Fecha Registro', 'Activo', 'Nota']
            ws.append([])  # Fila 4 vacia
            ws.append(encabezados)  # Fila 5
            
            # Estilos de encabezado
            fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            
            for col_num, encabezado in enumerate(encabezados, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = encabezado
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
            
            # Datos
            for idx, hab in enumerate(habitantes, start=6):
                ws.cell(row=idx, column=1).value = hab.get('folio', '')
                ws.cell(row=idx, column=2).value = hab.get('nombre', '')
                ws.cell(row=idx, column=3).value = hab.get('fecha_registro', '')
                ws.cell(row=idx, column=4).value = 'Si' if hab.get('activo') else 'No'
                ws.cell(row=idx, column=5).value = hab.get('nota', '')
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 30
            
            # Agregar estadisticas
            fila_stats = len(habitantes) + 7
            ws.cell(row=fila_stats, column=1).value = "Total Habitantes:"
            ws.cell(row=fila_stats, column=2).value = len(habitantes)
            
            activos = sum(1 for h in habitantes if h.get('activo'))
            ws.cell(row=fila_stats + 1, column=1).value = "Activos:"
            ws.cell(row=fila_stats + 1, column=2).value = activos
            
            # Guardar
            nombre_archivo = f"Reporte_Censo_{self._obtener_timestamp()}.xlsx"
            ruta = os.path.join(self.carpeta_reportes, nombre_archivo)
            
            wb.save(ruta)
            return ruta
        
        except ImportError:
            raise ImportError("openpyxl no esta instalado. Instala con: pip install openpyxl")
    
    # ========================================================================
    # REPORTES DE PAGOS
    # ========================================================================
    
    def generar_reporte_pagos_excel(self, pagos, habitantes, datos_usuario="Sistema"):
        """
        Genera reporte de pagos en Excel
        
        Args:
            pagos (list): Lista de pagos
            habitantes (dict): Diccionario con folio como clave
            datos_usuario (str): Usuario que genera el reporte
            
        Returns:
            str: Ruta del archivo generado
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Pagos"
            
            # Titulo
            ws['A1'] = "REPORTE DE PAGOS"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            
            # Encabezados
            encabezados = ['ID Pago', 'Habitante', 'Monto', 'Fecha', 'Estado']
            ws.append([])
            ws.append(encabezados)
            
            # Estilos
            fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            
            for col_num, encabezado in enumerate(encabezados, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = encabezado
                cell.fill = fill
                cell.font = font
            
            # Datos
            total_pagos = 0
            for idx, pago in enumerate(pagos, start=6):
                ws.cell(row=idx, column=1).value = pago.get('id', '')
                
                # Obtener nombre habitante
                folio = pago.get('habitante_folio', '')
                nombre = habitantes.get(folio, {}).get('nombre', 'N/A') if isinstance(habitantes, dict) else 'N/A'
                ws.cell(row=idx, column=2).value = nombre
                
                monto = pago.get('monto', 0)
                ws.cell(row=idx, column=3).value = monto
                ws.cell(row=idx, column=3).number_format = '$#,##0.00'
                
                ws.cell(row=idx, column=4).value = pago.get('fecha', '')
                ws.cell(row=idx, column=5).value = pago.get('estado', '')
                
                total_pagos += monto
            
            # Resumen
            fila_resumen = len(pagos) + 7
            ws.cell(row=fila_resumen, column=1).value = "TOTAL RECAUDADO:"
            ws.cell(row=fila_resumen, column=2).value = total_pagos
            ws.cell(row=fila_resumen, column=2).number_format = '$#,##0.00'
            ws.cell(row=fila_resumen, column=2).font = Font(bold=True)
            
            # Ancho de columnas
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            
            # Guardar
            nombre_archivo = f"Reporte_Pagos_{self._obtener_timestamp()}.xlsx"
            ruta = os.path.join(self.carpeta_reportes, nombre_archivo)
            wb.save(ruta)
            
            return ruta
        
        except ImportError:
            raise ImportError("openpyxl no esta instalado. Instala con: pip install openpyxl")
    
    # ========================================================================
    # REPORTES DE FAENAS
    # ========================================================================
    
    def generar_reporte_faenas_excel(self, faenas, habitantes, datos_usuario="Sistema"):
        """
        Genera reporte de faenas en Excel
        
        Args:
            faenas (list): Lista de faenas
            habitantes (list): Lista de habitantes
            datos_usuario (str): Usuario que genera el reporte
            
        Returns:
            str: Ruta del archivo generado
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Faenas"
            
            # Titulo
            ws['A1'] = "REPORTE DE FAENAS"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            
            # Encabezados
            encabezados = ['ID Faena', 'Fecha', 'Tipo', 'Participantes', 'Horas']
            ws.append([])
            ws.append(encabezados)
            
            # Estilos
            fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            font = Font(bold=True, color="FFFFFF")
            
            for col_num, encabezado in enumerate(encabezados, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = encabezado
                cell.fill = fill
                cell.font = font
            
            # Datos
            total_horas = 0
            for idx, faena in enumerate(faenas, start=6):
                ws.cell(row=idx, column=1).value = faena.get('id', '')
                ws.cell(row=idx, column=2).value = faena.get('fecha', '')
                ws.cell(row=idx, column=3).value = faena.get('tipo', '')
                
                participantes = faena.get('participantes', [])
                ws.cell(row=idx, column=4).value = len(participantes)
                
                horas = faena.get('horas', 0)
                ws.cell(row=idx, column=5).value = horas
                
                total_horas += horas
            
            # Resumen
            fila_resumen = len(faenas) + 7
            ws.cell(row=fila_resumen, column=1).value = "TOTAL HORAS:"
            ws.cell(row=fila_resumen, column=2).value = total_horas
            ws.cell(row=fila_resumen, column=2).font = Font(bold=True)
            
            ws.cell(row=fila_resumen + 1, column=1).value = "TOTAL FAENAS:"
            ws.cell(row=fila_resumen + 1, column=2).value = len(faenas)
            ws.cell(row=fila_resumen + 1, column=2).font = Font(bold=True)
            
            # Ancho
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 10
            
            # Guardar
            nombre_archivo = f"Reporte_Faenas_{self._obtener_timestamp()}.xlsx"
            ruta = os.path.join(self.carpeta_reportes, nombre_archivo)
            wb.save(ruta)
            
            return ruta
        
        except ImportError:
            raise ImportError("openpyxl no esta instalado. Instala con: pip install openpyxl")
    
    # ========================================================================
    # UTILIDADES
    # ========================================================================
    
    def obtener_reportes_generados(self):
        """
        Obtener lista de reportes generados
        
        Returns:
            list: Lista de archivos de reporte
        """
        if not os.path.exists(self.carpeta_reportes):
            return []
        
        reportes = []
        for archivo in os.listdir(self.carpeta_reportes):
            if archivo.endswith('.xlsx') or archivo.endswith('.pdf'):
                ruta_completa = os.path.join(self.carpeta_reportes, archivo)
                tamaño = os.path.getsize(ruta_completa)
                fecha_mod = datetime.fromtimestamp(os.path.getmtime(ruta_completa))
                
                reportes.append({
                    'nombre': archivo,
                    'ruta': ruta_completa,
                    'tamaño': tamaño,
                    'fecha_modificacion': fecha_mod.strftime('%d/%m/%Y %H:%M:%S')
                })
        
        return sorted(reportes, key=lambda x: x['fecha_modificacion'], reverse=True)
    
    def eliminar_reporte(self, nombre_archivo):
        """
        Eliminar un reporte
        
        Args:
            nombre_archivo (str): Nombre del archivo a eliminar
            
        Returns:
            bool: True si se elimino exitosamente
        """
        ruta = os.path.join(self.carpeta_reportes, nombre_archivo)
        
        if os.path.exists(ruta):
            os.remove(ruta)
            return True
        
        return False
    
    def limpiar_reportes_antiguos(self, dias=30):
        """
        Limpiar reportes mas antiguos a X dias
        
        Args:
            dias (int): Dias a mantener
            
        Returns:
            int: Cantidad de reportes eliminados
        """
        from datetime import timedelta
        
        fecha_limite = datetime.now() - timedelta(days=dias)
        eliminados = 0
        
        for archivo in os.listdir(self.carpeta_reportes):
            ruta = os.path.join(self.carpeta_reportes, archivo)
            fecha_mod = datetime.fromtimestamp(os.path.getmtime(ruta))
            
            if fecha_mod < fecha_limite:
                os.remove(ruta)
                eliminados += 1
        
        return eliminados


# Instancia global
_generador = GeneradorReportes()


def generar_reporte_censo(habitantes, usuario="Sistema"):
    """Generar reporte de censo"""
    return _generador.generar_reporte_censo_excel(habitantes, usuario)


def generar_reporte_pagos(pagos, habitantes, usuario="Sistema"):
    """Generar reporte de pagos"""
    return _generador.generar_reporte_pagos_excel(pagos, habitantes, usuario)


def generar_reporte_faenas(faenas, habitantes, usuario="Sistema"):
    """Generar reporte de faenas"""
    return _generador.generar_reporte_faenas_excel(faenas, habitantes, usuario)


def obtener_reportes():
    """Obtener lista de reportes generados"""
    return _generador.obtener_reportes_generados()
